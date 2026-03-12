"""
샘플 엑셀 파일 생성 스크립트
에이전트 테스트용 대용량 데이터가 담긴 엑셀 파일을 생성합니다.
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

PRODUCTS = ["노트북", "스마트폰", "태블릿", "모니터", "키보드", "마우스", "헤드셋", "웹캠", "프린터", "스캐너"]
REGIONS = ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "경기", "강원", "충북"]
SALESPEOPLE = [f"영업사원{i:02d}" for i in range(1, 21)]
CATEGORIES = {"노트북": "컴퓨터", "스마트폰": "모바일", "태블릿": "모바일",
              "모니터": "컴퓨터", "키보드": "주변기기", "마우스": "주변기기",
              "헤드셋": "음향기기", "웹캠": "주변기기", "프린터": "사무기기", "스캐너": "사무기기"}
BASE_PRICES = {"노트북": 1200000, "스마트폰": 800000, "태블릿": 600000,
               "모니터": 400000, "키보드": 80000, "마우스": 50000,
               "헤드셋": 150000, "웹캠": 100000, "프린터": 300000, "스캐너": 200000}


def generate_sales_data(n_rows: int = 5000) -> pd.DataFrame:
    random.seed(42)
    start_date = datetime(2023, 1, 1)

    records = []
    for i in range(n_rows):
        product = random.choice(PRODUCTS)
        base_price = BASE_PRICES[product]
        quantity = random.randint(1, 50)
        discount = random.choice([0, 0, 0, 0.05, 0.10, 0.15, 0.20])
        unit_price = int(base_price * (1 - discount))
        total = unit_price * quantity

        records.append({
            "주문ID": f"ORD-{i+1:05d}",
            "날짜": start_date + timedelta(days=random.randint(0, 364)),
            "제품명": product,
            "카테고리": CATEGORIES[product],
            "지역": random.choice(REGIONS),
            "담당자": random.choice(SALESPEOPLE),
            "수량": quantity,
            "단가": unit_price,
            "할인율": discount,
            "매출액": total,
            "반품여부": random.random() < 0.05,
        })

    df = pd.DataFrame(records)
    df["날짜"] = pd.to_datetime(df["날짜"])
    df = df.sort_values("날짜").reset_index(drop=True)
    return df


def generate_employee_data(n_rows: int = 200) -> pd.DataFrame:
    random.seed(123)
    departments = ["영업부", "개발부", "마케팅부", "인사부", "재무부", "운영부"]
    positions = ["사원", "대리", "과장", "차장", "부장", "이사"]
    position_salary = {"사원": 3000000, "대리": 3800000, "과장": 4800000,
                       "차장": 6000000, "부장": 7500000, "이사": 10000000}

    records = []
    for i in range(n_rows):
        position = random.choice(positions)
        base_salary = position_salary[position]
        records.append({
            "사원ID": f"EMP-{i+1:04d}",
            "이름": f"직원{i+1:03d}",
            "부서": random.choice(departments),
            "직급": position,
            "입사년도": random.randint(2010, 2023),
            "기본급": base_salary,
            "성과급": int(base_salary * random.uniform(0, 0.3)),
            "연간_총보수": int(base_salary * 12 * random.uniform(0.9, 1.2)),
            "평가등급": random.choice(["S", "A", "B", "C", "D"]),
            "재직여부": random.random() > 0.1,
        })

    return pd.DataFrame(records)


def generate_inventory_data(n_rows: int = 1000) -> pd.DataFrame:
    random.seed(456)
    warehouses = ["서울창고", "부산창고", "대구창고", "인천창고"]
    statuses = ["정상", "정상", "정상", "부족", "과잉", "단종예정"]

    records = []
    for i in range(n_rows):
        product = random.choice(PRODUCTS)
        records.append({
            "SKU": f"SKU-{i+1:05d}",
            "제품명": product,
            "카테고리": CATEGORIES[product],
            "창고": random.choice(warehouses),
            "재고수량": random.randint(0, 500),
            "최소재고": random.randint(10, 50),
            "최대재고": random.randint(200, 500),
            "단가": BASE_PRICES[product],
            "재고가치": BASE_PRICES[product] * random.randint(0, 500),
            "상태": random.choice(statuses),
            "최종입고일": datetime(2024, 1, 1) + timedelta(days=random.randint(0, 364)),
        })

    return pd.DataFrame(records)


def create_excel_with_summary(output_path: str):
    print(f"샘플 데이터 생성 중...")
    sales_df = generate_sales_data(5000)
    employee_df = generate_employee_data(200)
    inventory_df = generate_inventory_data(1000)

    wb = Workbook()

    # --- 매출 데이터 시트 ---
    ws_sales = wb.active
    ws_sales.title = "매출데이터"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for r in dataframe_to_rows(sales_df, index=False, header=True):
        ws_sales.append(r)

    for cell in ws_sales[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws_sales.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_sales.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    ws_sales.freeze_panes = "A2"

    # --- 요약 시트 (Excel 수식 사용) ---
    ws_summary = wb.create_sheet("요약분석")
    ws_summary["A1"] = "매출 요약 대시보드"
    ws_summary["A1"].font = Font(size=14, bold=True)

    summary_labels = [
        ("A3", "총 주문건수"),
        ("A4", "총 매출액"),
        ("A5", "평균 주문금액"),
        ("A6", "최대 단건 매출"),
        ("A7", "총 판매수량"),
        ("A8", "반품건수"),
        ("A9", "반품률 (%)"),
    ]
    summary_formulas = [
        ("B3", "=COUNTA(매출데이터!A2:A10000)"),
        ("B4", "=SUM(매출데이터!J2:J10000)"),
        ("B5", "=AVERAGE(매출데이터!J2:J10000)"),
        ("B6", "=MAX(매출데이터!J2:J10000)"),
        ("B7", "=SUM(매출데이터!G2:G10000)"),
        ("B8", '=COUNTIF(매출데이터!K2:K10000,TRUE)'),
        ("B9", "=B8/B3*100"),
    ]

    for cell_ref, label in summary_labels:
        ws_summary[cell_ref] = label
        ws_summary[cell_ref].font = Font(bold=True)

    for cell_ref, formula in summary_formulas:
        ws_summary[cell_ref] = formula

    ws_summary["A11"] = "카테고리별 매출"
    ws_summary["A11"].font = Font(size=12, bold=True)

    categories = list(set(CATEGORIES.values()))
    ws_summary["A12"] = "카테고리"
    ws_summary["B12"] = "매출액합계"
    ws_summary["C12"] = "주문건수"
    for h in ["A12", "B12", "C12"]:
        ws_summary[h].font = Font(bold=True)
        ws_summary[h].fill = PatternFill("solid", start_color="D9E1F2")

    for i, cat in enumerate(sorted(categories), start=13):
        ws_summary[f"A{i}"] = cat
        ws_summary[f"B{i}"] = f'=SUMIF(매출데이터!D:D,A{i},매출데이터!J:J)'
        ws_summary[f"C{i}"] = f'=COUNTIF(매출데이터!D:D,A{i})'

    for col in ["A", "B", "C"]:
        ws_summary.column_dimensions[col].width = 20

    # --- 직원 데이터 시트 ---
    ws_emp = wb.create_sheet("직원데이터")
    for r in dataframe_to_rows(employee_df, index=False, header=True):
        ws_emp.append(r)

    for cell in ws_emp[1]:
        cell.fill = PatternFill("solid", start_color="375623")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws_emp.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_emp.column_dimensions[col[0].column_letter].width = min(max_len + 2, 20)

    ws_emp.freeze_panes = "A2"

    # --- 재고 데이터 시트 ---
    ws_inv = wb.create_sheet("재고데이터")
    for r in dataframe_to_rows(inventory_df, index=False, header=True):
        ws_inv.append(r)

    for cell in ws_inv[1]:
        cell.fill = PatternFill("solid", start_color="7B3F00")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws_inv.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_inv.column_dimensions[col[0].column_letter].width = min(max_len + 2, 20)

    ws_inv.freeze_panes = "A2"

    wb.save(output_path)
    print(f"샘플 파일 저장 완료: {output_path}")
    print(f"  - 매출데이터: {len(sales_df):,}행")
    print(f"  - 직원데이터: {len(employee_df):,}행")
    print(f"  - 재고데이터: {len(inventory_df):,}행")
    print(f"  - 요약분석: Excel 수식 기반 대시보드")


def main():
    output_path = DATA_DIR / "sample_business_data.xlsx"
    create_excel_with_summary(str(output_path))

    print("\n테스트 질문 예시:")
    print("  - '전체 매출 현황을 분석해줘'")
    print("  - '어떤 제품이 가장 많이 팔렸어?'")
    print("  - '지역별 매출을 비교해줘'")
    print("  - '2023년 월별 매출 트렌드를 알려줘'")
    print("  - '부서별 평균 연봉을 계산해줘'")
    print("  - '재고가 부족한 제품 목록을 뽑아줘'")


if __name__ == "__main__":
    main()
